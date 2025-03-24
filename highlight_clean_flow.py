import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

# ---------- Utility Functions ----------
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def highlight_and_flag_totals(file_bytes):
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    flagged_cells = []

    for ws in wb.worksheets:
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                value = str(cell.value) if cell.value else ""
                if "Total" in value and "(" in value and ")" in value:
                    cell.fill = yellow_fill
                    flagged_cells.append((ws.title, cell.coordinate, value))

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream, flagged_cells


def clean_flagged_totals(file_bytes):
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)

    def remove_parentheses_content(text):
        return re.sub(r'\s*\([^)]*\)', '', text).strip()

    for ws in wb.worksheets:
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                if cell.value and "Total" in str(cell.value):
                    cell.value = remove_parentheses_content(str(cell.value))
                    cell.fill = None

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream


# ---------- Streamlit App Flow ----------
st.set_page_config(page_title="Filing Cleanup Wizard", layout="centered")
st.title("🧾 Filing Cleanup Wizard")

if "step" not in st.session_state:
    st.session_state.step = 1
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "flagged_cells" not in st.session_state:
    st.session_state.flagged_cells = []

# Step 1: Upload Excel
if st.session_state.step == 1:
    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
    if uploaded_file:
        file_bytes = uploaded_file.read()
        highlighted_file, flagged = highlight_and_flag_totals(file_bytes)
        st.session_state.excel_bytes = highlighted_file
        st.session_state.flagged_cells = flagged

        st.success(f"Found {len(flagged)} potentially incorrect 'Total' cells.")
        st.session_state.step = 2
        st.experimental_rerun()

# Step 2: Show flagged cells
elif st.session_state.step == 2:
    st.subheader("🔍 Review Flagged Cells")
    if st.session_state.flagged_cells:
        for sheet, coord, val in st.session_state.flagged_cells:
            st.write(f"- **{sheet}**!{coord} → `{val}`")

        if st.button("Yes, clean these cells"):
            cleaned_file = clean_flagged_totals(st.session_state.excel_bytes)
            st.session_state.excel_bytes = cleaned_file
            st.session_state.step = 3
            st.experimental_rerun()

        if st.button("No, leave them as-is"):
            st.session_state.step = 3
            st.experimental_rerun()
    else:
        st.info("No problematic 'Total' cells found. Skipping ahead.")
        st.session_state.step = 3
        st.experimental_rerun()

# Step 3: Download
elif st.session_state.step == 3:
    st.subheader("✅ Final Step: Download Cleaned File")
    st.download_button(
        label="Download Cleaned Excel",
        data=st.session_state.excel_bytes,
        file_name="cleaned_filing.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    if st.button("Start Over"):
        for key in ["step", "excel_bytes", "flagged_cells"]:
            st.session_state.pop(key, None)
        st.experimental_rerun()
