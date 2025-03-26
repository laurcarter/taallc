import openpyxl
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string

def create_summary(focus_ws, max_row):
    summary_row = 8  # Starting row for the summary
    
    # Loop through the Focus sheet to find and summarize subtotals
    for row in range(8, max_row + 1):
        c_value = focus_ws.cell(row=row, column=3).value  # Value in column C
        d_value = focus_ws.cell(row=row, column=4).value  # Value in column D

        # Check if the value in column C contains the word "Total"
        if c_value and "Total" in str(c_value):
            # Split the value in column C into number ID and "Total"
            focus_ws.cell(row=summary_row, column=9).value = str(c_value).replace("Total", "").strip()  # Number ID in column I
            focus_ws.cell(row=summary_row, column=10).value = "Total"  # Word "Total" in column J
            focus_ws.cell(row=summary_row, column=11).value = d_value  # Subtotal value in column K
            summary_row += 1  # Move to the next row for the next summary item


def apply_income_expense_totals_ssoi(ssoi_ws, max_row):
    # Call the categorize_income_expense_ssoi function to get the income and expense sums
    income_sum, expense_sum = categorize_income_expense_ssoi(ssoi_ws, max_row)


    # Initialize row trackers for income and expense sections
    income_rows = []
    expense_rows = []

    # Loop through the rows again to identify income and expense sections (rows with green and red fills)
    for row_idx in range(8, max_row + 1):
        c_value = ssoi_ws.cell(row=row_idx, column=3).value  # Column C
        d_value = ssoi_ws.cell(row=row_idx, column=4).value  # Column D

        # Only consider rows with income or expense values (skip others)
        if c_value is None or d_value is None:
            continue

        # Check for income or expense rows based on the value in column C
        numeric_value = ''.join(filter(str.isdigit, str(c_value)))
        if numeric_value.isdigit():
            numeric_value = float(numeric_value)

            # Add the row index to the respective list (income or expense) based on the value in column C
            if numeric_value <= 11:
                income_rows.append(row_idx)
            else:
                expense_rows.append(row_idx)

    # Insert the income sum into the last row of the income section
    if income_rows:
        last_income_row = income_rows[-1]  # Get the last row in the income section
        ssoi_ws.cell(row=last_income_row, column=6).value = round(income_sum, 2)  # Insert sum into column F
        ssoi_ws.cell(row=last_income_row, column=6).font = Font(bold=True)  # Make it bold

    # Insert the expense sum into the last row of the expense section
    if expense_rows:
        last_expense_row = expense_rows[-1]  # Get the last row in the expense section
        ssoi_ws.cell(row=last_expense_row, column=6).value = round(expense_sum, 2)  # Insert sum into column F
        ssoi_ws.cell(row=last_expense_row, column=6).font = Font(bold=True)  # Make it bold

    # Calculate the result by subtracting expenses from income
    result = income_sum - expense_sum

    # Find the last used row in column C to determine where to place the "NET INCOME" value
    last_row = max_row
    for row in range(max_row, 7, -1):  # Start from max_row and move upwards
        if ssoi_ws.cell(row=row, column=3).value is not None:
            last_row = row
            break

    # Place the result in the cell below the last used row in column F
    ssoi_ws.cell(row=last_row + 1, column=5).value = "NET INCOME"  # Column E for "NET INCOME"
    ssoi_ws.cell(row=last_row + 1, column=5).font = Font(bold=True)  # Make the "NET INCOME" bold

    # Place the result in column F
    ssoi_ws.cell(row=last_row + 1, column=6).value = round(result, 2)  # Column F for result
    ssoi_ws.cell(row=last_row + 1, column=6).font = Font(bold=True)  # Make the result bold


def categorize_income_expense_ssoi(ssoi_ws, max_row):
    # Define the colors for income and expense
    light_green_fill = PatternFill(start_color="D9F2D1", end_color="D9F2D1", fill_type="solid")  # Light green
    light_red_fill = PatternFill(start_color="F9E2D2", end_color="F9E2D2", fill_type="solid")  # Light red

    income_sum = 0
    expense_sum = 0

    # Loop through each row starting from row 8
    for row in range(8, max_row + 1):
        c_value = ssoi_ws.cell(row=row, column=3).value  # Value in column C
        d_value = ssoi_ws.cell(row=row, column=4).value  # Value in column D

        # If there's no value in column C, skip this row
        if c_value is None or c_value == "":
            continue

        # Scrub letters from the value in column C (take only the numeric part)
        numeric_value = ''.join(filter(str.isdigit, str(c_value)))

        # Ensure the numeric value is treated as a number
        if numeric_value.isdigit():
            numeric_value = float(numeric_value)

            # Categorize as income (less than or equal to 11) or expense (greater than 11)
            if numeric_value <= 11:
                # Income: Add the value in column D to the income sum and apply green color
                if isinstance(d_value, (int, float)):
                    income_sum += d_value / 2  # Divide by 2 as per the given logic
                    
                    # Apply green color to income rows (Columns C to F)
                    for col in range(3, 7):  # Columns C to F
                        ssoi_ws.cell(row=row, column=col).fill = light_green_fill

            elif numeric_value > 11:
                # Expense: Add the value in column D to the expense sum and apply red color
                if isinstance(d_value, (int, float)):
                    expense_sum += d_value / 2  # Divide by 2 as per the given logic
                    
                    # Apply red color to expense rows (Columns C to F)
                    for col in range(3, 7):  # Columns C to F
                        ssoi_ws.cell(row=row, column=col).fill = light_red_fill

    return income_sum, expense_sum

    return income_sum, expense_sum


def apply_income_expense_totals(focus_ws, max_row):
    # Call the categorize_income_expense function to get the income and expense sums
    income_sum, expense_sum = categorize_income_expense(focus_ws, max_row)

    # Divide the income and expense sums by 2
    income_sum /= 2
    expense_sum /= 2

    # Initialize row trackers for income and expense sections
    income_rows = []
    expense_rows = []

    # Loop through the rows again to identify income and expense sections (rows with green and red fills)
    for row_idx in range(8, max_row + 1):
        c_value = focus_ws.cell(row=row_idx, column=3).value  # Column C
        d_value = focus_ws.cell(row=row_idx, column=4).value  # Column D

        # Only consider rows with income or expense values (skip others)
        if c_value is None or d_value is None:
            continue

        # Check for income or expense rows based on the value in column C
        numeric_value = ''.join(filter(str.isdigit, str(c_value)))
        if numeric_value.isdigit():
            numeric_value = float(numeric_value)

            # Add the row index to the respective list (income or expense) based on the value in column C
            if numeric_value < 4000:
                income_rows.append(row_idx)
            else:
                expense_rows.append(row_idx)

    # Insert the income sum into the last row of the income section
    if income_rows:
        last_income_row = income_rows[-1]  # Get the last row in the income section
        focus_ws.cell(row=last_income_row, column=6).value = round(income_sum, 2)  # Insert sum into column F
        focus_ws.cell(row=last_income_row, column=6).font = Font(bold=True)  # Make it bold

    # Insert the expense sum into the last row of the expense section
    if expense_rows:
        last_expense_row = expense_rows[-1]  # Get the last row in the expense section
        focus_ws.cell(row=last_expense_row, column=6).value = round(expense_sum, 2)  # Insert sum into column F
        focus_ws.cell(row=last_expense_row, column=6).font = Font(bold=True)  # Make it bold

    # Calculate the result by subtracting expenses from income
    result = income_sum - expense_sum

    # Find the last used row in column C to determine where to place the "NET INCOME" value
    last_row = max_row
    for row in range(max_row, 7, -1):  # Start from max_row and move upwards
        if focus_ws.cell(row=row, column=3).value is not None:
            last_row = row
            break

    # Place the result in the cell below the last used row in column F
    focus_ws.cell(row=last_row + 1, column=5).value = "NET INCOME"  # Column E for "NET INCOME"
    focus_ws.cell(row=last_row + 1, column=5).font = Font(bold=True)  # Make the "NET INCOME" bold

    # Place the result in column F
    focus_ws.cell(row=last_row + 1, column=6).value = round(result, 2)  # Column F for result
    focus_ws.cell(row=last_row + 1, column=6).font = Font(bold=True)  # Make the result bold



def categorize_income_expense(focus_ws, max_row):
    # Define the colors
    light_green_fill = PatternFill(start_color="D9F2D1", end_color="D9F2D1", fill_type="solid")  # Light green
    light_red_fill = PatternFill(start_color="F9E2D2", end_color="F9E2D2", fill_type="solid")  # Light red

    income_sum = 0
    expense_sum = 0

    # Loop through each row starting from row 8
    for row in range(8, max_row + 1):
        c_value = focus_ws.cell(row=row, column=3).value  # Value in column C
        d_value = focus_ws.cell(row=row, column=4).value  # Value in column D

        # If there's no value in column C, skip this row
        if c_value is None or c_value == "":
            continue

        # Scrub letters from the value in column C (take only the numeric part)
        numeric_value = ''.join(filter(str.isdigit, str(c_value)))

        # Ensure the numeric value is treated as a number
        if numeric_value.isdigit():
            numeric_value = float(numeric_value)

            # Categorize as income (less than 4000) or expense (greater than or equal to 4000)
            if numeric_value < 4000:
                # Apply color for income (light green)
                for col in range(3, 7):  # Columns C to F
                    focus_ws.cell(row=row, column=col).fill = light_green_fill

                # Add the value from column D to the income sum
                if isinstance(d_value, (int, float)):
                    income_sum += d_value
            else:
                # Apply color for expense (light red)
                for col in range(3, 7):  # Columns C to F
                    focus_ws.cell(row=row, column=col).fill = light_red_fill

                # Add the value from column D to the expense sum
                if isinstance(d_value, (int, float)):
                    expense_sum += d_value

    return income_sum, expense_sum




def delete_blank_rows(ws, max_row):
    # Start from row 8 and go downwards
    row_idx = 8

    while row_idx <= max_row:
        c_value = ws.cell(row=row_idx, column=3).value  # Get the value from column C
        d_value = ws.cell(row=row_idx, column=4).value  # Get the value from column D

        # If both columns C and D are blank, delete the row
        if (c_value is None or c_value == "") and (d_value is None or d_value == ""):
            ws.delete_rows(row_idx)  # Delete the current row
            max_row -= 1  # Decrease max_row because we just deleted a row
        else:
            row_idx += 1  # Move to the next row if not both columns are blank


def apply_subtotals(focus_ws, ssoi_ws, max_row):
    # Apply subtotals to both the Focus and SSOI sheets
    apply_subtotals_for_sheet(focus_ws, max_row)
    apply_subtotals_for_sheet(ssoi_ws, max_row)

def apply_subtotals_for_sheet(ws, max_row):
    current_value = None  # Tracks the current group in column C
    total_sum = 0  # Tracks the sum of the values in column D for the current group
    last_row = max_row  # The last row in the sheet

    row_idx = 8  # Start from row 8 as per your description

    while row_idx <= max_row:
        c_value = ws.cell(row=row_idx, column=3).value  # Get the value in column C
        d_value = ws.cell(row=row_idx, column=4).value  # Get the value in column D

        # If the value in column C is empty, skip this row
        if c_value is None or c_value == "":
            row_idx += 1
            continue

        # If the current value in column C is different from the previous one, insert a subtotal row
        if c_value != current_value:
            # If we have already encountered a group, insert the total row
            if current_value is not None:
                ws.insert_rows(row_idx)
                ws.cell(row=row_idx, column=3).value = f"{current_value} Total"  # Insert "Total" in column C
                ws.cell(row=row_idx, column=4).value = total_sum  # Insert the sum in column D

                # Move the row index down because we just inserted a new row
                row_idx += 1

            # Reset for the new group
            current_value = c_value
            total_sum = d_value if isinstance(d_value, (int, float)) else 0  # Start sum with first value of the new group
        else:
            # Add the current value in column D to the running total
            if isinstance(d_value, (int, float)):
                total_sum += d_value

        row_idx += 1  # Move to the next row

    # Handle the last group after the loop ends (ensure the final group is processed)
    if current_value is not None:
        ws.insert_rows(last_row + 1)  # Add the subtotal row at the end
        ws.cell(row=last_row + 1, column=3).value = f"{current_value} Total"
        ws.cell(row=last_row + 1, column=4).value = total_sum


def clean_ss01_column(ssoi_ws, max_row):
    # Loop through each cell in column C of the SSOI sheet starting from row 5
    for row in range(5, max_row + 1):
        cell = ssoi_ws.cell(row=row, column=3)
        cell_value = str(cell.value)  # Ensure cell value is a string
        
        # Check if the value starts with an apostrophe and then handle it
        if cell_value.startswith("'"):
            # Remove the apostrophe and any leading zeros
            cleaned_value = cell_value.lstrip("'0")
            
            # If the cleaned value is empty, set the value to '0', otherwise just the cleaned value
            if cleaned_value == "":
                cleaned_value = "0"
            
            cell.value = cleaned_value  # Update the cell with the cleaned value


# Sorting function for SSOI sheet column C
def sort_ssoi_sheet(ssoi_ws, max_row):
    # Loop through column C starting from row 5 to find and delete rows with empty cells
    for row in range(max_row, 4, -1):  # Go from the bottom to the top to avoid skipping rows
        cell = ssoi_ws.cell(row=row, column=3)
        if cell.value is None or cell.value == "":  # If the cell in column C is empty
            ssoi_ws.delete_rows(row)  # Delete the row

    # Create a list to hold rows with their corresponding values in column C
    rows = []

    # Loop through column C starting from row 5
    for row in range(5, max_row + 1):
        cell = ssoi_ws.cell(row=row, column=3)
        if cell.value is not None:
            c_value = str(cell.value).strip()
            
            # Append the entire row with its value in column C
            rows.append((row, c_value, [ssoi_ws.cell(row=row, column=col).value for col in range(1, ssoi_ws.max_column + 1)]))

    # Sort rows based on the value in column C (ascending)
    # First, sort numeric values (no letters), then alphanumeric
    rows.sort(key=lambda x: (int(x[1]) if x[1].isdigit() else float('inf'), x[1]))

    # Clear the existing values in the sheet starting from row 5
    for row in range(5, max_row + 1):
        for col in range(1, ssoi_ws.max_column + 1):
            ssoi_ws.cell(row=row, column=col).value = None

    # Write the sorted rows back into the sheet
    for idx, (original_row, _, row_values) in enumerate(rows, start=5):
        for col_idx, value in enumerate(row_values, start=1):
            ssoi_ws.cell(row=idx, column=col_idx).value = value


            
def sort_focus_sheet(focus_ws, max_row):
    # Step 1: Delete rows with empty values in column C
    for row in range(max_row, 7, -1):  # Start from the bottom to avoid skipping rows
        cell = focus_ws.cell(row=row, column=3)
        if cell.value is None or cell.value == "":  # If the cell in column C is empty
            focus_ws.delete_rows(row)  # Delete the entire row

    # Step 2: Create a list to hold rows with their corresponding values in column C
    rows = []

    # Step 3: Loop through column C starting from row 8
    for row in range(8, max_row + 1):
        cell = focus_ws.cell(row=row, column=3)
        if cell.value is not None:
            c_value = str(cell.value).strip()
            
            # Append the entire row with its value in column C
            rows.append((row, c_value, [focus_ws.cell(row=row, column=col).value for col in range(1, focus_ws.max_column + 1)]))

    # Step 4: Sort rows based on the value in column C (ascending)
    # First, sort numeric values (no letters), then alphanumeric
    rows.sort(key=lambda x: (int(x[1]) if x[1].isdigit() else float('inf'), x[1]))

    # Step 5: Clear the existing values in the Focus sheet starting from row 8
    for row in range(8, max_row + 1):
        for col in range(1, focus_ws.max_column + 1):
            focus_ws.cell(row=row, column=col).value = None

    # Step 6: Write the sorted rows back into the Focus sheet
    for idx, (original_row, _, row_values) in enumerate(rows, start=8):
        for col_idx, value in enumerate(row_values, start=1):
            focus_ws.cell(row=idx, column=col_idx).value = value

def secondary_sort_ssoi_sheet(ssoi_ws, max_row):
    # Create lists to hold rows with numeric values in column C and those with alphanumeric values
    numeric_rows = []
    alphanumeric_rows = []

    # Loop through column C starting from row 8
    for row in range(8, max_row + 1):
        c_value = ssoi_ws.cell(row=row, column=3).value
        d_value = ssoi_ws.cell(row=row, column=4).value
        
        # Skip rows where column C is empty
        if c_value is None or c_value == "":
            continue
        
        # Check if the value in column D is numeric, otherwise treat it as the lowest value
        if isinstance(d_value, (int, float)):
            d_value = float(d_value)  # Ensure the value is treated as a float
        else:
            d_value = float('-inf')  # Non-numeric values will be treated as the lowest possible

        # Separate numeric values (without letters) and alphanumeric values (with letters)
        if str(c_value).replace('.', '', 1).isdigit():  # Check if it's a number (allowing for decimal points)
            numeric_rows.append((row, c_value, d_value, [ssoi_ws.cell(row=row, column=col).value for col in range(1, ssoi_ws.max_column + 1)]))
        else:
            alphanumeric_rows.append((row, c_value, d_value, [ssoi_ws.cell(row=row, column=col).value for col in range(1, ssoi_ws.max_column + 1)]))

    # Sort the numeric rows in ascending order based on column C
    numeric_rows.sort(key=lambda x: x[1])

    # Sort the alphanumeric rows in ascending order based on column C and descending for D values
    alphanumeric_rows.sort(key=lambda x: (x[1], -x[2]))

    # Combine numeric rows and alphanumeric rows
    all_sorted_rows = numeric_rows + alphanumeric_rows

    # Write the sorted rows back to the sheet
    new_row_idx = 8  # Start writing from row 8
    for _, _, _, row_values in all_sorted_rows:
        for col_idx, value in enumerate(row_values, start=1):
            ssoi_ws.cell(row=new_row_idx, column=col_idx).value = value
        new_row_idx += 1

def secondary_sort_focus_sheet(focus_ws, max_row):
    # Create a list to hold rows with their corresponding values from columns C and D
    rows = []

    # Loop through column C starting from row 8 (instead of 5)
    for row in range(8, max_row + 1):
        c_value = focus_ws.cell(row=row, column=3).value
        d_value = focus_ws.cell(row=row, column=4).value
        
        # Skip rows where column C is empty
        if c_value is None or c_value == "":
            continue

        # Ensure column D is treated as a numeric value if possible
        if isinstance(d_value, (int, float)):
            d_value = float(d_value)  # Ensure the value is treated as a float
        else:
            d_value = float('-inf')  # Non-numeric values will be treated as the lowest possible
        
        # Append the entire row along with values from column C and D
        rows.append((row, c_value, d_value, [focus_ws.cell(row=row, column=col).value for col in range(1, focus_ws.max_column + 1)]))

    # Sort rows based on column C (ascending) and column D (descending for same values in C)
    rows.sort(key=lambda x: (x[1], -x[2]) if x[1] is not None else ("", float('inf')))

    # Clear the existing values in the sheet starting from row 8
    for row in range(8, max_row + 1):
        for col in range(1, focus_ws.max_column + 1):
            focus_ws.cell(row=row, column=col).value = None

    # Write the sorted rows back into the sheet starting from row 8
    new_row_idx = 8
    for _, _, _, row_values in rows:
        for col_idx, value in enumerate(row_values, start=1):
            focus_ws.cell(row=new_row_idx, column=col_idx).value = value
        new_row_idx += 1



def run_full_pl_macro(file_bytes):
    # Ensure file_bytes is a BytesIO object
    if isinstance(file_bytes, BytesIO):
        file_obj = file_bytes
    else:
        file_obj = BytesIO(file_bytes)

    wb = load_workbook(filename=file_obj)
    ws = wb.active  # Get the active worksheet

    # Create Focus and SSOI worksheets
    focus_ws = wb.create_sheet(title="Focus")
    ssoi_ws = wb.create_sheet(title="SSOI")

    # Find the last used row and column in the original sheet
    max_row = ws.max_row
    max_col = ws.max_column

    # Copy content from original sheet to Focus and SSOI as plain values
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            focus_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            ssoi_ws.cell(row=cell.row, column=cell.column, value=cell.value)

    # Format IDs in column C of SSOI sheet as text with leading zeros
    for row in range(1, max_row + 1):
        cell = ssoi_ws.cell(row=row, column=3)
        val = str(cell.value).strip() if cell.value is not None else ""
        if len(val) == 1:
            cell.value = f"0{val}"  # Add leading zero for single-character
        elif len(val) >= 2:
            cell.value = f"'{val}"  # Add apostrophe to treat as text

    # Insert a temporary column to format IDs as text in SSOI sheet (mimic the Insert Column behavior)
    ssoi_ws.insert_cols(4)

    # Apply the NumberFormat for column C in the SSOI sheet (mimic Excel's "@")
    ssoi_ws.column_dimensions['C'].number_format = '@'

    # *** NEW CODE START ***
    # Copy the formatted values back to column C and delete the temporary column D
    for row in range(1, max_row + 1):
        ssoi_ws.cell(row=row, column=3, value=ssoi_ws.cell(row=row, column=4).value)
    
    # Delete the temporary column D (column 4)
    ssoi_ws.delete_cols(4)
    
    # *** NEW CODE END ***

    # Delimit Column A by '(' and extract to Column B (both sheets)
    for sheet in [focus_ws, ssoi_ws]:
        for row in range(1, max_row + 1):
            val = sheet.cell(row=row, column=1).value
            if val and '(' in str(val):
                parts = str(val).split('(', 1)
                sheet.cell(row=row, column=1).value = parts[0].strip()
                sheet.cell(row=row, column=2).value = parts[1].strip()

    # Delimit Column B by '/' and process (Focus and SSOI)
    for row in range(1, max_row + 1):
        # Focus Sheet: Left part to B, Right part to C
        val = focus_ws.cell(row=row, column=2).value
        if val and '/' in str(val):
            parts = str(val).split('/')
            focus_ws.cell(row=row, column=2).value = parts[0].strip().replace("(", "")
            focus_ws.cell(row=row, column=3).value = parts[1].strip().replace(")", "").replace("/", "")

        # SSOI Sheet: Right part to B, Left part to C, strip ')' if present
        val = ssoi_ws.cell(row=row, column=2).value
        if val and '/' in str(val):
            parts = str(val).split('/')
            right = parts[1].strip().replace(")", "")
            left = parts[0].strip().replace("(", "")
            ssoi_ws.cell(row=row, column=2).value = right
            ssoi_ws.cell(row=row, column=3).value = left

    # Strip out the opening parenthesis in Column B for both sheets
    for sheet in [focus_ws, ssoi_ws]:
        for row in range(1, max_row + 1):
            val = sheet.cell(row=row, column=2).value
            if val:
                # Ensure val is a string before calling replace()
                val = str(val)
                sheet.cell(row=row, column=2).value = val.replace("(", "")
    
    # Strip out the slash and closing parenthesis in Column C for both sheets
    for sheet in [focus_ws, ssoi_ws]:
        for row in range(1, max_row + 1):
            val = sheet.cell(row=row, column=3).value
            if val:
                # Ensure val is a string before calling replace()
                val = str(val)
                sheet.cell(row=row, column=3).value = val.replace("/", "").replace(")", "")

    # Wipe out column B of the SSOI sheet
    for row in range(1, max_row + 1):
        ssoi_ws.cell(row=row, column=2).value = None
    
    # Copy Column C from Focus sheet to Column B in SSOI sheet
    for row in range(1, max_row + 1):
        focus_value = focus_ws.cell(row=row, column=3).value
        ssoi_ws.cell(row=row, column=2).value = focus_value

    # Wipe out column B of the Focus sheet
    for row in range(1, max_row + 1):
        focus_ws.cell(row=row, column=2).value = None
    
    # Copy Column C from SSOI sheet to Column B in Focus sheet
    for row in range(1, max_row + 1):
        ssoi_value = ssoi_ws.cell(row=row, column=3).value
        focus_ws.cell(row=row, column=2).value = ssoi_value

  # Clear column C in both Focus and SSOI sheets
    for row in range(1, max_row + 1):
        focus_ws.cell(row=row, column=3).value = None
        ssoi_ws.cell(row=row, column=3).value = None

    # Copy Column B from original sheet to Column D in both Focus and SSOI sheets
    for row in range(1, max_row + 1):
        original_value = ws.cell(row=row, column=2).value
        focus_ws.cell(row=row, column=4).value = original_value
        ssoi_ws.cell(row=row, column=4).value = original_value

    # Move the entire sheet over by two columns in both Focus and SSOI sheets
    focus_ws.insert_cols(1, 2)  # Insert two columns at the beginning of the Focus sheet
    ssoi_ws.insert_cols(1, 2)   # Insert two columns at the beginning of the SSOI sheet
    
    # Move data in columns starting from row 5 in the Focus sheet
    for row in range(5, max_row + 1):
        focus_ws.cell(row=row, column=3).offset(0, 2).value = focus_ws.cell(row=row, column=3).value
        focus_ws.cell(row=row, column=3).value = None  # Clear original cell
    
        focus_ws.cell(row=row, column=4).offset(0, -1).value = focus_ws.cell(row=row, column=4).value
        focus_ws.cell(row=row, column=4).value = None  # Clear original cell
    
        focus_ws.cell(row=row, column=6).offset(0, -2).value = focus_ws.cell(row=row, column=6).value
        focus_ws.cell(row=row, column=6).value = None  # Clear original cell
    
    # Same process for the SSOI sheet
    for row in range(5, max_row + 1):
        ssoi_ws.cell(row=row, column=3).offset(0, 2).value = ssoi_ws.cell(row=row, column=3).value
        ssoi_ws.cell(row=row, column=3).value = None  # Clear original cell
    
        ssoi_ws.cell(row=row, column=4).offset(0, -1).value = ssoi_ws.cell(row=row, column=4).value
        ssoi_ws.cell(row=row, column=4).value = None  # Clear original cell
    
        ssoi_ws.cell(row=row, column=6).offset(0, -2).value = ssoi_ws.cell(row=row, column=6).value
        ssoi_ws.cell(row=row, column=6).value = None  # Clear original cell
        
    # Clean the SSOI column C before sorting
    clean_ss01_column(ssoi_ws, max_row)

    # Now, call the sort function to sort the SSOI sheet based on column C
    sort_ssoi_sheet(ssoi_ws, max_row)
# new
    # Add column titles in row 4
    focus_ws["C4"] = "Focus"
    focus_ws["D4"] = "Amount"
    focus_ws["E4"] = "Description"
    focus_ws["F4"] = "Totals"
    
    ssoi_ws["C4"] = "SSOI"
    ssoi_ws["D4"] = "Amount"
    ssoi_ws["E4"] = "Description"
    ssoi_ws["F4"] = "Totals"
    
    # Shift everything below row 4 down by 3 rows in both sheets
    focus_ws.insert_rows(4, amount=3)  # Insert 3 rows at row 4 in focus_ws
    ssoi_ws.insert_rows(4, amount=3)   # Insert 3 rows at row 4 in ssoi_ws

    # Call the sort_focus_sheet function after the rest of the operations in the macro
    sort_focus_sheet(focus_ws, max_row)

    # After sorting column C (done by previous functions), call this function for secondary sorting
    secondary_sort_ssoi_sheet(ssoi_ws, max_row)
    secondary_sort_focus_sheet(focus_ws, max_row)

    # Define fill color (black) and font color (white)
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_font = Font(color="FFFFFF")

    # Fill columns C to F in row 7 with black and change the text color to white in the Focus sheet
    for col in ["C", "D", "E", "F"]:
        focus_ws[f"{col}7"].fill = black_fill
        focus_ws[f"{col}7"].font = white_font

    # Fill columns C to F in row 7 with black and change the text color to white in the SSOI sheet
    for col in ["C", "D", "E", "F"]:
        ssoi_ws[f"{col}7"].fill = black_fill
        ssoi_ws[f"{col}7"].font = white_font

    # Format column D and F in the Focus sheet to show numbers with thousand commas
    for row in range(1, max_row + 1):
        focus_ws.cell(row=row, column=4).number_format = "#,##0"  # Column D
        focus_ws.cell(row=row, column=6).number_format = "#,##0"  # Column F

    # Format column D and F in the SSOI sheet to show numbers with thousand commas
    for row in range(1, max_row + 1):
        ssoi_ws.cell(row=row, column=4).number_format = "#,##0"  # Column D
        ssoi_ws.cell(row=row, column=6).number_format = "#,##0"  # Column F

    # Increase the width of column E to double the default width in the Focus sheet
    focus_ws.column_dimensions["E"].width = focus_ws.column_dimensions["E"].width * 2.5

    # Increase the width of column E to double the default width in the SSOI sheet
    ssoi_ws.column_dimensions["E"].width = ssoi_ws.column_dimensions["E"].width * 2.5

    #subtotals
    apply_subtotals(focus_ws, ssoi_ws, max_row)
    
    # Call this function for both sheets
    delete_blank_rows(focus_ws, max_row)  # For Focus sheet
    delete_blank_rows(ssoi_ws, max_row)   # For SSOI sheet
    
    

    # You can now use income_sum and expense_sum in your further calculations
    apply_income_expense_totals(focus_ws, max_row)
    apply_income_expense_totals_ssoi(ssoi_ws, max_row)

    create_summary(focus_ws, max_row)


    # Ensure to save the workbook after sorting if needed
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream

