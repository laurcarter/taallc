import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.title("🟨 Highlight Total Cells with IDs in Excel")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

def highlight_cells(file):
    wb = load_workbook(filename=file, data_only=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for ws in wb.worksheets:
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                value = str(cell.value) if cell.value else ""
                if "Total" in value and "(" in value and ")" in value:
                    cell.fill = yellow_fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if uploaded_file:
    st.success("File uploaded successfully!")
    processed_file = highlight_cells(uploaded_file)
    st.download_button(
        label="Download Highlighted File",
        data=processed_file,
        file_name="highlighted_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
