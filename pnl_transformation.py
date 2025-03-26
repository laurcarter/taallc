import streamlit as st
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
import re

# ---------- Utility Functions ----------
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Function to highlight flagged totals
def highlight_and_flag_totals(file_bytes):
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    flagged_cells = []

    for ws in wb.worksheets:
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                value = str(cell.value) if cell.value else ""
                if "Total" in value and "(" in value and ")" in value:
                    cell.fill = yellow_fill
                    flagged_cells.append((ws.title, cell.coordinate, value))

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream, flagged_cells

# Function to clean flagged totals (remove parentheses and 'Total')
def clean_flagged_totals(file_bytes):
    if isinstance(file_bytes, BytesIO):
        file_obj = file_bytes
    else:
        file_obj = BytesIO(file_bytes)

    wb = load_workbook(filename=file_obj, data_only=True)

    def remove_parentheses_content(text):
        return re.sub(r'\s*\([^)]*\)', '', text).strip()

    for ws in wb.worksheets:
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                if cell.value and "Total" in str(cell.value):
                    cell.value = remove_parentheses_content(str(cell.value))
                    cell.fill = PatternFill()  # Reset to default empty fill

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream

# Function to perform P&L transformation
def perform_pnl_transformation(file_bytes):
    from pnl_macro_translation import run_full_pl_macro
    return run_full_pl_macro(file_bytes)

# ---------- Streamlit App Flow ----------
st.set_page_config(page_title="Filing Cleanup Wizard", layout="centered")
st.title("🧾 Filing Cleanup Wizard")

if "step" not in st.session_state:
    st.session_state.step = 1
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "flagged_cells" not in st.session_state:
    st.session_state.flagged_cells = []

# Step 1: Personal Information Collection
if st.session_state.step == 1:
    st.subheader("To get started, we'll need some information about you.")
    first_name = st.text_input("First Name")
    middle_initial = st.text_input("Middle Initial")
    last_name = st.text_input("Last Name")
    suffix = st.selectbox("Suffix (Jr., Sr., III)", ["", "Jr.", "Sr.", "III"])
    occupation = st.text_input("Occupation")
    employer = st.text_input("Employer")
    dob = st.date_input("Date of Birth", format="MM/dd/yyyy")
    phone_number = st.text_input("Phone Number")

    if st.button("Continue"):
        # Store the collected information in session state
        st.session_state.first_name = first_name
        st.session_state.middle_initial = middle_initial
        st.session_state.last_name = last_name
        st.session_state.suffix = suffix
        st.session_state.occupation = occupation
        st.session_state.employer = employer
        st.session_state.dob = dob
        st.session_state.phone_number = phone_number
        st.session_state.step = 2

# Step 2: Personal Information Summary
elif st.session_state.step == 2:
    st.subheader("🔍 Your Personal Information Summary")
    st.write(f"**First Name:** {st.session_state.first_name}")
    st.write(f"**Middle Initial:** {st.session_state.middle_initial}")
    st.write(f"**Last Name:** {st.session_state.last_name}")
    st.write(f"**Suffix:** {st.session_state.suffix}")
    st.write(f"**Occupation:** {st.session_state.occupation}")
    st.write(f"**Employer:** {st.session_state.employer}")
    st.write(f"**Date of Birth:** {st.session_state.dob.strftime('%m/%d/%Y')}")
    st.write(f"**Phone Number:** {st.session_state.phone_number}")

    st.write("### Phone Numbers:")
    phone_number_1 = st.text_input("Phone Number 1", value=st.session_state.phone_number)
    phone_number_2 = st.text_input("Phone Number 2")
    phone_number_3 = st.text_input("Phone Number 3")
    
    if st.button("Continue"):
        st.session_state.phone_number_1 = phone_number_1
        st.session_state.phone_number_2 = phone_number_2
        st.session_state.phone_number_3 = phone_number_3
        st.session_state.step = 3

# Step 3: Upload Excel File
elif st.session_state.step == 3:
    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
    if uploaded_file:
        file_bytes = uploaded_file.read()
        highlighted_file, flagged = highlight_and_flag_totals(file_bytes)
        st.session_state.excel_bytes = highlighted_file
        st.session_state.flagged_cells = flagged

        st.success(f"Found {len(flagged)} potentially incorrect 'Total' cells.")
        if st.button("Continue"):
            st.session_state.step = 4

# Step 4: Show flagged cells for review
elif st.session_state.step == 4:
    st.subheader("🔍 Review Flagged Cells")
    if st.session_state.flagged_cells:
        for sheet, coord, val in st.session_state.flagged_cells:
            st.write(f"- **{sheet}**!{coord} → `{val}`")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Yes, clean these cells"):
                cleaned_file = clean_flagged_totals(st.session_state.excel_bytes)
                st.session_state.excel_bytes = cleaned_file
                st.session_state.step = 5
        with col2:
            if st.button("No, leave them as-is"):
                st.session_state.step = 5

    else:
        st.info("No problematic 'Total' cells found. Skipping ahead.")
        if st.button("Continue"):
            st.session_state.step = 5

# Step 5: Choose Transformation Type
elif st.session_state.step == 5:
    st.subheader("🔧 What type of filing is this?")
    choice = st.radio("Select your filing type:", ["Profit & Loss (P&L)", "Balance Sheet"], index=0)
    if st.button("Run Transformation"):
        if choice == "Profit & Loss (P&L)":
            st.session_state.excel_bytes = perform_pnl_transformation(st.session_state.excel_bytes)
        st.session_state.step = 6

# Step 6: Download Final Processed File
elif st.session_state.step == 6:
    st.subheader("✅ Final Step: Download Processed File")
    st.download_button(
        label="Download Final Excel",
        data=st.session_state.excel_bytes,
        file_name="final_filing.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    if st.button("Start Over"):
        for key in ["step", "excel_bytes", "flagged_cells"]:
            st.session_state.pop(key, None)
