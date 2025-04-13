import streamlit as st
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
import re
from collapse import collapse_sheet
from efocus import efocus_focus  # Import the efocus logic
import pandas as pd


# ---------- Utility Functions ----------
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def check_and_prompt_for_net_income(focus_ws):
    # Look for "Net Income" in column A and check if it's coded with parentheses
    for row in range(8, focus_ws.max_row + 1):
        cell_value = str(focus_ws.cell(row=row, column=1).value).strip()

        if 'Net Income' in cell_value and '(' not in cell_value:
            # "Net Income" found without parentheses, ask for input
            net_income_input = st.text_input("Your 'Net Income' is not coded. Please provide the value:", "")
            
            if net_income_input:
                # If the user enters a value, update the cell with parentheses
                updated_value = f"Net Income ({net_income_input})"
                focus_ws.cell(row=row, column=1).value = updated_value
                st.session_state.excel_bytes = save_updated_excel(focus_ws)  # Save updated Excel file
                return True  # Indicates that an update was made

    return False  # No update made


def save_updated_excel(focus_ws):
    output = BytesIO()  # Create a new BytesIO object
    wb = focus_ws.parent  # Get the parent workbook of the active sheet
    wb.save(output)  # Save the workbook to the BytesIO object
    output.seek(0)  # Reset the cursor position in the BytesIO object
    return output  # Return the updated file as BytesIO





# Function to highlight flagged totals
def highlight_and_flag_totals(file_bytes):
    # Ensure file_bytes is wrapped in BytesIO before passing to load_workbook
    if not isinstance(file_bytes, BytesIO):
        file_bytes = BytesIO(file_bytes)

    wb = load_workbook(filename=file_bytes, data_only=True)
    flagged_cells = []

    for ws in wb.worksheets:
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                value = str(cell.value) if cell.value else ""
                if "Total" in value and "(" in value and ")" in value:
                    cell.fill = yellow_fill
                    flagged_cells.append((ws.title, cell.coordinate, value))

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream, flagged_cells


# Function to clean flagged totals (remove parentheses and 'Total')
def clean_flagged_totals(file_bytes):
    if isinstance(file_bytes, BytesIO):
        file_obj = file_bytes
    else:
        file_obj = BytesIO(file_bytes)

    wb = load_workbook(filename=file_obj, data_only=True)

    def remove_parentheses_content(text):
        return re.sub(r'\s*\([^)]*\)', '', text).strip()

    for ws in wb.worksheets:
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                if cell.value and "Total" in str(cell.value):
                    cell.value = remove_parentheses_content(str(cell.value))  # Clean text
                    cell.fill = PatternFill()  # Reset the highlight fill

    # Save the updated file and return as BytesIO
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream



# Function to perform P&L transformation
def perform_pnl_transformation(file_bytes):
    from pnl_macro_translation import run_full_pl_macro
    return run_full_pl_macro(file_bytes)


# Function to perform Balance transformation
def perform_balance_transformation(file_bytes):
    from balance import balance_focus_grouping 
    return balance_focus_grouping(file_bytes)

# ---------- Streamlit App Flow ----------
st.set_page_config(page_title="Personal Information Collection", layout="wide")

# Step Handling
if "step" not in st.session_state:
    st.session_state.step = 1
if "personal_info" not in st.session_state:
    st.session_state.personal_info = {}

# Step 1: Personal Information Collection
if st.session_state.step == 1:
    st.title("🧾 Personal Information Collection")  # Title for Step 1
    st.write("To get started, we'll need some information about you.")  # Description for Step 1

    with st.expander("Step 1: Tell us about yourself", expanded=True):
        first_name = st.text_input("First Name")
        middle_initial = st.text_input("Middle Initial")
        last_name = st.text_input("Last Name")
        suffix = st.selectbox("Suffix (Jr., Sr., III)", ["", "Jr.", "Sr.", "III"])
        occupation = st.text_input("Occupation")
        employer = st.text_input("Employer")
        dob = st.date_input("Date of Birth")  # Streamlit will handle the date format automatically
        phone_number = st.text_input("Phone Number")

        if st.button("Continue"):
            # Store the collected information in session state
            st.session_state.personal_info = {
                "first_name": first_name,
                "middle_initial": middle_initial,
                "last_name": last_name,
                "suffix": suffix,
                "occupation": occupation,
                "employer": employer,
                "dob": dob,
                "phone_number": phone_number
            }
            st.session_state.step = 2

# Step 2: Personal Information Summary
elif st.session_state.step == 2:
    st.title("🔍 Review Your Personal Information")  # Title for Step 2
    st.write("Please review your personal information before continuing.")  # Description for Step 2

    with st.expander("Step 2: Review Your Information", expanded=True):
        st.write(f"**First Name:** {st.session_state.personal_info['first_name']}")
        st.write(f"**Middle Initial:** {st.session_state.personal_info['middle_initial']}")
        st.write(f"**Last Name:** {st.session_state.personal_info['last_name']}")
        st.write(f"**Suffix:** {st.session_state.personal_info['suffix']}")
        st.write(f"**Occupation:** {st.session_state.personal_info['occupation']}")
        st.write(f"**Employer:** {st.session_state.personal_info['employer']}")
        st.write(f"**Date of Birth:** {st.session_state.personal_info['dob'].strftime('%m/%d/%Y')}")
        st.write(f"**Phone Number:** {st.session_state.personal_info['phone_number']}")

        if st.button("Continue"):
            st.session_state.step = 3

# Step 3: Upload Excel File
elif st.session_state.step == 3:
    st.title("📂 Upload Your Excel File")  # Title for Step 3
    st.write("Please upload your Excel file for auto-filing.")  # Description for Step 3

    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
    if uploaded_file:
        file_bytes = uploaded_file.read()

        # Load the workbook to check the number of sheets
        wb = load_workbook(filename=BytesIO(file_bytes))
        sheet_names = wb.sheetnames  # List of sheet names
        if len(sheet_names) > 1:
            # If multiple sheets are detected, ask the user to resubmit a new file with a single sheet
            st.error("Your file contains multiple sheets. Please resubmit a new file with only one sheet.")
            st.session_state.step = 3  # Stay on Step 3 so they can upload again
        else:
            # If there is only one sheet, proceed with the regular logic
            sheet = wb.active

            # Now highlight and flag totals on the selected sheet
            highlighted_file, flagged = highlight_and_flag_totals(file_bytes)
            st.session_state.excel_bytes = highlighted_file  # Store the highlighted file in session state
            st.session_state.flagged_cells = flagged  # Store the flagged cells
    
            st.success(f"Found {len(flagged)} potentially incorrect 'Total' cells.")
    
            # Continue button to move to the next step
            if st.button("Continue"):
                    st.session_state.step = 4  # Move to Step 4





# Step 4: Show flagged cells for review and clean if needed
elif st.session_state.step == 4:
    st.title("🔍 Review Flagged Cells")  # Title for Step 4
    st.write("Review and clean any flagged cells, or leave them as-is.")  # Description for Step 4

    # Ensure flagged_cells exist in session_state
    if 'flagged_cells' not in st.session_state:
        st.session_state.flagged_cells = []

    if st.session_state.flagged_cells:
        # Display all the flagged cells with their coordinates
        for sheet, coord, val in st.session_state.flagged_cells:
            st.write(f"- **{sheet}**!{coord} → {val}")

        col1, col2 = st.columns(2)

        # Button for cleaning flagged totals
        with col1:
            if st.button("Yes, clean these cells"):
                # Get the current file from session state
                file_bytes = st.session_state.excel_bytes
                
                # Clean the flagged totals (this function will modify the file)
                cleaned_file = clean_flagged_totals(file_bytes)

                # Update the session state with the cleaned file
                st.session_state.excel_bytes = cleaned_file  # Store the cleaned file in session state

                # Proceed to Step 5 after cleaning
                st.session_state.step = 5  # Move to Step 5 (next step)

        # Button for leaving the flagged totals as-is (just keep them highlighted)
        with col2:
            if st.button("No, leave them as-is"):
                # Ensure we keep the file as-is with highlighted totals, no cleaning
                file_bytes = st.session_state.excel_bytes
                
                # Just highlight the totals again and move to Step 5
                highlighted_file, _ = highlight_and_flag_totals(file_bytes)  # Re-highlight if needed
                st.session_state.excel_bytes = highlighted_file  # Store the highlighted file in session state

                # Proceed to Step 5 without cleaning
                st.session_state.step = 5  # Move to Step 5 (next step)

    else:
        # If no flagged cells, display a message
        st.info("No problematic 'Total' cells found. Skipping ahead.")
        if st.button("Continue"):
            # Proceed to Step 5 if no flagged cells
            st.session_state.step = 5  # Skip to Step 5 if no flagged cells



# Step 5: Choose Transformation Type
elif st.session_state.step == 5:
    file_bytes = st.session_state.excel_bytes  # The current file in session state

    # Load the workbook
    if not isinstance(file_bytes, BytesIO):
        file_bytes = BytesIO(file_bytes)

    wb = load_workbook(file_bytes)
    sheet = wb.active  # Use the active sheet from the loaded workbook
    focus_ws = wb.active  # Get the active sheet from the loaded workbook
    
    # Check for empty cells in column A (rows 5-10) and collapse if necessary
    empty_cell_count = 0
    total_cells_to_check = 6  # Checking rows 5 to 10 (6 rows total)
    for row in range(5, 11):  # Rows 5 to 10 in column A
        cell = sheet.cell(row=row, column=1)
        if cell.value is None or str(cell.value).strip() == "":
            empty_cell_count += 1
    if empty_cell_count == total_cells_to_check:
        collapsed_file = collapse_sheet(file_bytes)
        if isinstance(collapsed_file, BytesIO):  # Ensure it returns BytesIO
            st.session_state.excel_bytes = collapsed_file
        else:
            st.error("Error: Collapse function did not return a valid file.")
    
    st.title("🔧 What type of filing is this?")
    st.write("Select the type of filing for this document.")

    # Radio button for selecting filing type
    choice = st.radio("Select your filing type:", ["Profit & Loss (P&L)", "Balance Sheet"], index=0)

    # Create a flag to indicate if Net Income needs to be updated (only relevant for Balance Sheet)
    net_income_updated = False  # Default value for Net Income update status

    # Only check for Net Income when Balance Sheet is selected
    if choice == "Balance Sheet":
        # If Net Income is not coded, prompt the user
        net_income_updated = check_and_prompt_for_net_income(sheet)
        
        if net_income_updated:
            st.success("Net Income has been updated!")  # Let the user know the update was made

    # "Run Transformation" button
    if st.button("Run Transformation"):
        # Only run transformations when the button is clicked
        if choice == "Profit & Loss (P&L)":
            # Proceed with P&L transformation
            st.session_state.excel_bytes = perform_pnl_transformation(st.session_state.excel_bytes)
        
        elif choice == "Balance Sheet":
            # Run balance transformation only after Net Income update (if applicable)
            st.session_state.excel_bytes = perform_balance_transformation(st.session_state.excel_bytes)
        st.session_state.step = 6

# Step 6: Final Processed File Download
elif st.session_state.step == 6:
    st.title("✅ Final Step: Download Processed File")  # Title for Step 6
    st.write("Download the final processed file.")  # Description for Step 6

    # Provide download button for the processed file
    st.download_button(
        label="Download Final Excel",
        data=st.session_state.excel_bytes,
        file_name="final_filing.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Button to start over and reset session state
    if st.button("Start Over"):
        for key in ["step", "excel_bytes", "flagged_cells"]:
            st.session_state.pop(key, None)

    # Button to continue to eFocus creation (Step 7)
    if st.button("Continue to eFocus creation"):
        # Move to Step 7
        st.session_state.step = 7



# Step 7: eFocus Creation (Use the file from Step 6)
elif st.session_state.step == 7:
    file_bytes = st.session_state.excel_bytes  # The current file in session state
    st.title("📂 eFocus Creation")  # Title for Step 7
    st.write("Continue to create eFocus using the uploaded Focus file and select the client.")  # Description for Step 7

    # Streamlit UI for the file upload and processing
    client_data_file = st.file_uploader("Upload the Client Data file", type=["xlsx"])

    if client_data_file:
        # Read the files as bytes
        client_data_bytes = client_data_file.read()

        # Process the files with the efocus_focus function and capture selected client
        transformed_file, selected_client = efocus_focus(file_bytes, client_data_bytes)

        if transformed_file:
            # Store the transformed file in session state
            st.session_state.excel_bytes = transformed_file
        
            # Reset the BytesIO object to the beginning before download
            st.session_state.excel_bytes.seek(0)
        
            # Use the selected client's name in the file name
            file_name = f"efocus_{selected_client}.xlsx"  # Client name added to the file name

            # Move to Step 8 after processing the file and client selection
            st.session_state.step = 8  # Move to Step 8 (with client questions)

        else:
            st.info("Please select a client name to proceed.")
    
    # "Continue" button for moving to Step 8
    if st.button("Continue to Step 8"):
        st.session_state.step = 8  # Proceed to Step 8



# Step 8: Client Information and Questions (New step after eFocus creation)
elif st.session_state.step == 8:
    st.title("📝 Client Information and Questions")  # Title for Step 8
    st.write("Please answer the following questions to proceed with the next steps.")  # Description for Step 8

    # Question 1: Ask if it's a monthly or quarterly filing
    filing_frequency = st.radio("Is this a monthly or quarterly filing?", ["Monthly", "Quarterly"])

    # Store the filing frequency directly in session state
    st.session_state.filing_frequency = filing_frequency  # Assign directly to session_state variable

    # Conditionally display further questions based on the filing frequency
    if filing_frequency == "Monthly":
        # Ask if this is a P&L for monthly income purposes
        is_pnl = st.radio("For monthly income purposes: Is this filing from a P&L?", ["Yes", "No"])
        st.session_state.is_pnl = is_pnl  # Save the P&L response in session state
        
        # If the answer is Yes, ask for the monthly income amount
        if is_pnl == "Yes":
            monthly_income = st.number_input("Please enter the monthly income amount:", min_value=0, step=1000)
            st.session_state.monthly_income = monthly_income  # Save the income amount

    # After answering, the user can click "Continue" to move on to the next step
    if st.button("Continue to Next Step"):
        # Store answers to all questions in session state (as shown above)
        st.session_state.client_answers = {
            "filing_frequency": st.session_state.filing_frequency,
            "is_pnl": st.session_state.is_pnl,
            "monthly_income": st.session_state.monthly_income if 'monthly_income' in st.session_state else None
        }

        # Proceed to the next step after answering questions (adjust step as needed)
        st.session_state.step = 9  # Move to Step 9 (or whatever comes next)
# Step 9: Ending Equity Balance (Ask for the ending equity balance)
elif st.session_state.step == 9:
    st.title("💰 Ending Equity Balance")  # Title for Step 9
    st.write("Please enter the ending equity balance from the previous filing period.")  # Description

    # Ask for the ending equity balance
    ending_equity_balance = st.number_input("Enter the ending equity balance:", min_value=0.0, step=1000.0)
    
    # Save the value in session state
    st.session_state.ending_equity_balance = ending_equity_balance  # Assign to session_state variable

    # Show a confirmation message with the entered value
    st.write(f"You entered: ${ending_equity_balance:.2f}")

    # Continue to next step (if needed)
    if st.button("Continue to Next Step"):
        # Proceed to the next step after answering the question
        st.session_state.step = 10  # Move to Step 10 (or whatever comes next)

