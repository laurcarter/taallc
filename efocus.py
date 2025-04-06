import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def match_and_copy_values(focus_ws, focus_target_ws):
    # Loop through rows 8 to 40 in column I of the Focus sheet
    for row in range(8, 41):  # Rows 8 to 40 (inclusive)
        focus_value = focus_ws.cell(row=row, column=9).value  # Value in column I of Focus
        
        # If the cell has a value
        if focus_value:
            # Strip any "I" or leading zeros from the Focus value
            focus_value_stripped = str(focus_value).lstrip("I0").strip()
            
            # Now search for this stripped value in FocusTarget column A
            for target_row in range(1, focus_target_ws.max_row + 1):
                target_value = str(focus_target_ws.cell(row=target_row, column=1).value).lstrip("I0").strip()
                
                # If a match is found, get the value from column J in Focus and paste it in column B of FocusTarget
                if focus_value_stripped == target_value:
                    focus_value_j = focus_ws.cell(row=row, column=10).value  # Get value from column J of Focus
                    focus_target_ws.cell(row=target_row, column=2, value=focus_value_j)  # Paste it in column B of FocusTarget
                    break  # Exit loop once a match is found



def efocus_focus(file_bytes, client_data_bytes):
    # Ensure the Focus file bytes are wrapped in BytesIO if not already
    if not isinstance(file_bytes, BytesIO):
        file_bytes = BytesIO(file_bytes)  # Wrap Focus file as BytesIO if not already

    # Reset the file stream to the start before passing to openpyxl
    file_bytes.seek(0)

    # Load the Focus sheet from the uploaded file (file_bytes)
    try:
        wb = load_workbook(filename=file_bytes)
        focus_ws = wb['Focus']  # Assuming the Focus sheet is already available
    except Exception as e:
        # Handle potential errors related to loading the workbook
        print(f"Error loading Excel file: {e}")
        st.error(f"Error loading Excel file: {e}")
        return None, None

    # Handle the Client Data bytes, wrap in BytesIO if necessary
    client_data_bytes_io = BytesIO(client_data_bytes)  # Wrap Client Data file as BytesIO if not already
    client_data_bytes_io.seek(0)  # Reset the stream to the start

    # Load the client data from the second uploaded file (client_data_bytes)
    client_data = pd.read_excel(client_data_bytes_io, header=None)  # Reading client data without headers

    client_names = []
    for col in range(2, client_data.shape[1], 2):
        cell_value = str(client_data.iloc[0, col]).strip()
        if cell_value and 'Unnamed' not in cell_value:
            client_names.append(cell_value)

    # If no valid client names were found, show a message and exit
    if not client_names:
        st.error("No valid client names found in the client data.")
        return None, None  # Return None for both if no client names are found

    # Display the valid client names as clickable buttons in columns
    selected_client = None
    columns = st.columns(4)  # Create 4 columns to stack the buttons

    # Loop through client names and place them into columns
    for idx, client in enumerate(client_names):
        col_idx = idx % 4  # Determine the column index based on the position
        if columns[col_idx].button(client):
            selected_client = client  # Store the selected client name when the button is clicked

    # If a client has been selected, proceed
    if selected_client:
        st.write(f"You selected: {selected_client}")

        # Find the column index of the selected client name in the client data
        client_column = client_names.index(selected_client) * 2 + 3  # Adjust column index for 0-indexing (C is column 3)
        st.write(f"Client Column: {client_column}")  # For debugging purposes

        # Create the "FocusTarget" sheet in the original workbook
        focus_target_ws = wb.create_sheet(title="FocusTarget")

        # Copy column A from the client data file (rows 1 to 275) into "FocusTarget"
        client_column_a = client_data.iloc[:, 0]  # Column A (no header)
        for i, value in enumerate(client_column_a, start=1):
            focus_target_ws.cell(row=i, column=1, value=value)

        # Copy the selected client column from the client data (rows 1 to 275) into "FocusTarget"
        client_column_data = client_data.iloc[:, client_column - 1]  # Selected column based on the client name
        for i, value in enumerate(client_column_data, start=1):
            focus_target_ws.cell(row=i, column=2, value=value)

        # Set header for the new column B
        focus_target_ws.cell(row=1, column=2, value=selected_client)

        # Paste column B from client data into "FocusTarget" column C
        client_column_b = client_data.iloc[:, 1]  # Column B
        for i, value in enumerate(client_column_b, start=1):
            focus_target_ws.cell(row=i, column=3, value=value)

        # Now, copy the cell from Row 1, Column B into Row 4, Column E in FocusTarget
        client_data_b1 = focus_target_ws.cell(row=1, column=2).value  # Get the value from Row 1, Column B of FocusTarget
        focus_target_ws.cell(row=4, column=5, value=client_data_b1)  # Paste it into Row 4, Column E of FocusTarget

        # Insert "FOCUS" into row 4, column G in all caps and make it bold
        focus_target_ws.cell(row=4, column=7, value="FOCUS")  # Insert "FOCUS" into column G
        focus_target_ws.cell(row=4, column=5).font = Font(bold=True)  # Make cell in column E bold
        focus_target_ws.cell(row=4, column=7).font = Font(bold=True)  # Make cell in column G bold

        # Add "Item Value" in row 1, column B
        focus_target_ws.cell(row=1, column=2, value="Item Value")

        # Bold all of row 1
        for cell in focus_target_ws[1]:
            cell.font = Font(bold=True)

        # Call the function to process Focus and FocusTarget
        match_and_copy_values(focus_ws, focus_target_ws)

        # Rename the FocusTarget sheet
        focus_target_ws.title = "Filing Items Focus"

        # Save the modified workbook to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Move the cursor to the beginning of the BytesIO object

        # Return both the transformed file and selected client
        return output, selected_client

    # If no client has been selected yet, inform the user
    st.info("Please select a client name to proceed.")
    return None, None  # Return None if no client is selected
