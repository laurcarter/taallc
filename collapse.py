from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def collapse_sheet(file_bytes):
    wb = load_workbook(filename=BytesIO(file_bytes))  # Ensure file_bytes is wrapped in BytesIO
    ws = wb.active  # Active sheet

    # Create a temporary sheet to hold the data
    temp_ws = wb.create_sheet("TempSheet")
    
    # Copy the content of the active sheet to the temporary sheet
    for row in ws.iter_rows():
        for cell in row:
            temp_ws[cell.coordinate].value = cell.value

    # Step 1: Find the first occurrence of the word "Assets" in column A
    asset_found = False
    for i, row in enumerate(temp_ws.iter_rows(min_col=1, max_col=1, max_row=temp_ws.max_row), 1):
        if row[0].value and isinstance(row[0].value, str) and "assets" in row[0].value.lower():
            asset_found = True
            asset_row = i
            break
    
    # If "Assets" is found, remove all rows above it
    if asset_found:
        if asset_row > 2:
            temp_ws.delete_rows(1, asset_row - 1)

    # Step 2: Find the first non-blank row and column by checking left to right, top to bottom
    first_data_row = None
    first_data_col = None
    for row in temp_ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                first_data_row = cell.row
                first_data_col = cell.column
                break
        if first_data_row is not None:
            break

    if first_data_row is None:
        raise ValueError("No data found in the sheet.")

    # Get the last column used in the sheet
    last_col = temp_ws.max_column

    # Delete all columns to the right of the first data column
    if first_data_col < last_col:
        temp_ws.delete_cols(first_data_col + 1, last_col - first_data_col)

    # Step 3: Create the CleanedSheet for the final output
    clean_ws = wb.create_sheet("CleanedSheet")

    # Loop through each row and collapse the account names into one column
    for i, row in enumerate(temp_ws.iter_rows(min_row=first_data_row, max_row=temp_ws.max_row), 1):
        account_names = ""
        has_balance = False

        # Loop through the row to concatenate account names (exclude last column for balance)
        for j in range(1, len(row)):
            if row[j].value:
                account_names += f" {row[j].value}"

        # Trim leading/trailing spaces
        account_names = account_names.strip()

        # Place the concatenated account names in the new sheet (column A)
        clean_ws.cell(row=i, column=1).value = account_names

        # Check if the last column has a balance (numeric value or formatted as currency)
        if isinstance(row[-1].value, (int, float)) and row[-1].value != "":
            has_balance = True

        # If there's a balance, move it to column B in the cleaned sheet
        if has_balance:
            clean_ws.cell(row=i, column=2).value = row[-1].value

    # Step 4: Delete the temporary sheet
    wb.remove(temp_ws)

    # Step 5: Insert a row at the top of the CleanedSheet for headers
    clean_ws.insert_rows(1)
    clean_ws.cell(row=1, column=1).value = "Account Names"
    clean_ws.cell(row=1, column=2).value = "Balance"

    # Move the CleanedSheet to the front of the workbook using the `move_sheet` method
    wb.move_sheet(clean_ws, offset=-len(wb.sheetnames))  # Move the CleanedSheet to the front

    # Save the workbook and return the processed data
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    return output_stream
