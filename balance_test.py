import openpyxl
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO


def apply_comma_format_no_decimal(focus_ws, start_row=8, end_row=100):
    # Create a style with number format for comma separation (no decimals)
    comma_style = NamedStyle(name="comma_style", number_format="#,##0")

    # Apply the style to each cell in Column F (from start_row to end_row)
    for row in range(start_row, end_row + 1):
        focus_ws.cell(row=row, column=6).style = comma_style  # Column F is column 6


def calculate_totals(focus_ws, start_row=8, end_row=100):
    # Step 1: Initialize variables for totals
    total_assets = 0
    total_liabilities = 0
    total_equity = 0

    # Step 2: Loop through each row and sum the amounts for each group (based on the numeric code in Column C)
    for row in range(start_row, end_row + 1):
        c_value = str(focus_ws.cell(row=row, column=3).value).strip()  # Value in Column C
        try:
            # Extract numeric code, skipping "Total" rows
            if "Total" in c_value:
                numeric_code = int(c_value.replace("Total", "").strip())
            else:
                numeric_code = int(c_value)
        except ValueError:
            numeric_code = None
        
        # Step 3: Add amounts to the correct total group
        if numeric_code is not None:
            # Assets (200 to 940)
            if 200 <= numeric_code <= 940:
                amount = focus_ws.cell(row=row, column=4).value
                if isinstance(amount, (int, float)):
                    total_assets += amount
            # Liabilities (970 to 1760)
            elif 970 <= numeric_code <= 1760:
                amount = focus_ws.cell(row=row, column=4).value
                if isinstance(amount, (int, float)):
                    total_liabilities += amount
            # Ownership Equity (1020 or 1770 to 1810)
            elif numeric_code == 1020 or (1770 <= numeric_code <= 1810):
                amount = focus_ws.cell(row=row, column=4).value
                if isinstance(amount, (int, float)):
                    total_equity += amount

    # Step 4: Insert the totals in the last row of each group (in Column F)
    # Assets: Find the last row of Assets and insert the total
    for row in range(end_row, start_row - 1, -1):
        c_value = str(focus_ws.cell(row=row, column=3).value).strip()
        try:
            numeric_code = int(c_value.replace("Total", "").strip()) if "Total" in c_value else int(c_value)
        except ValueError:
            numeric_code = None
        
        if numeric_code is not None and 200 <= numeric_code <= 940:  # Assets
            focus_ws.cell(row=row, column=6).value = total_assets / 2  # Insert total in Column F
            focus_ws.cell(row=row, column=6).font = openpyxl.styles.Font(bold=True)  # Bold the total
            break  # Exit after updating the last row of Assets

    # Liabilities: Find the last row of Liabilities and insert the total
    for row in range(end_row, start_row - 1, -1):
        c_value = str(focus_ws.cell(row=row, column=3).value).strip()
        try:
            numeric_code = int(c_value.replace("Total", "").strip()) if "Total" in c_value else int(c_value)
        except ValueError:
            numeric_code = None
        
        if numeric_code is not None and 970 <= numeric_code <= 1760:  # Liabilities
            focus_ws.cell(row=row, column=6).value = total_liabilities / 2  # Insert total in Column F
            focus_ws.cell(row=row, column=6).font = openpyxl.styles.Font(bold=True)  # Bold the total
            break  # Exit after updating the last row of Liabilities

    # Ownership Equity: Find the last row of Ownership Equity and insert the total
    for row in range(end_row, start_row - 1, -1):
        c_value = str(focus_ws.cell(row=row, column=3).value).strip()
        try:
            numeric_code = int(c_value.replace("Total", "").strip()) if "Total" in c_value else int(c_value)
        except ValueError:
            numeric_code = None
        
        if numeric_code is not None and (numeric_code == 1020 or 1770 <= numeric_code <= 1810):  # Ownership Equity
            focus_ws.cell(row=row, column=6).value = total_equity / 2  # Insert total in Column F
            focus_ws.cell(row=row, column=6).font = openpyxl.styles.Font(bold=True)  # Bold the total
            break  # Exit after updating the last row of Ownership Equity


def apply_color_coding(focus_ws, start_row=8):
    # Step 1: Get the last row with data in Column C
    max_row = focus_ws.max_row
    
    # Initialize color fills for each category
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Assets (green)
    orange_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Liabilities (orange)
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Ownership Equity (blue)
    
    # Step 2: Loop through each row starting from row 8
    for row in range(start_row, max_row + 1):
        c_value = str(focus_ws.cell(row=row, column=3).value).strip()  # Value in Column C
        
        # Step 3: Check if the cell contains a numeric code or a subtotal
        try:
            if c_value.isdigit():  # If it's a pure numeric value
                numeric_code = int(c_value)
            elif "Total" in c_value and c_value.replace("Total", "").strip().isdigit():  # Subtotal row
                numeric_code = int(c_value.replace("Total", "").strip())
            else:
                numeric_code = None  # Skip non-numeric and non-subtotal rows
        except ValueError:
            numeric_code = None  # If there's an error in conversion, skip that row
        
        # Step 4: Apply color-coding based on the numeric codes
        if numeric_code is not None:
            if 200 <= numeric_code <= 940:  # Assets
                fill = green_fill
            elif 970 <= numeric_code <= 1760:  # Liabilities
                fill = orange_fill
            elif numeric_code == 1020 or (1770 <= numeric_code <= 1810):  # Ownership Equity
                fill = blue_fill
            else:
                fill = None  # Default: No color for other codes

            # Step 5: Apply color fill to columns C to F
            if fill:
                for col in range(3, 7):  # Columns C to F
                    focus_ws.cell(row=row, column=col).fill = fill

def move_last_total_below_group(focus_ws, start_row=8, end_row=100):
    # Step 1: Find the last group of rows and the corresponding "Total" row
    current_group = None
    last_row_in_group = None
    total_row = None
    last_total_row = None

    # Iterate through rows to find the last group and its corresponding "Total"
    for row in range(start_row, end_row + 1):
        c_value = focus_ws.cell(row=row, column=3).value  # Value in Column C
        
        if c_value and "Total" in str(c_value):  # Found a "Total" row
            total_row = row  # Track the "Total" row
            last_total_row = total_row  # Keep updating to the latest "Total" row
        elif c_value and c_value != "Total":
            # If it's a non-"Total" row, update the group and the last row
            current_group = c_value
            last_row_in_group = row

    # Step 2: Move the last "Total" row below its corresponding group
    if last_total_row is not None:
        # Move the "Total" row to right below the last non-"Total" row in the group
        focus_ws.cell(row=last_row_in_group + 1, column=3).value = focus_ws.cell(row=last_total_row, column=3).value
        focus_ws.cell(row=last_row_in_group + 1, column=4).value = focus_ws.cell(row=last_total_row, column=4).value
        focus_ws.cell(row=last_row_in_group + 1, column=5).value = focus_ws.cell(row=last_total_row, column=5).value
        
        # Clear the current "Total" row
        for col in range(1, focus_ws.max_column + 1):
            focus_ws.cell(row=last_total_row, column=col).value = None

        # Update the last row to reflect the moved "Total"
        last_row_in_group += 1




def apply_random_formatting(focus_ws, max_row):
    # Comma formatting and rounding for Focus sheet only

    # Column J (now column 10 after deletion)
    for row in range(8, max_row + 1):
        cell = focus_ws.cell(row=row, column=10)  # Column J
        if isinstance(cell.value, (int, float)):
            # Round to 0 decimal places and apply comma format
            cell.value = round(cell.value, 0)
            cell.number_format = "#,##0"  # Apply comma style formatting

    # Column F (column 6)
    for row in range(8, max_row + 1):
        cell = focus_ws.cell(row=row, column=6)  # Column F
        if isinstance(cell.value, (int, float)):
            # Apply comma style formatting
            cell.number_format = "#,##0"

    # Column D (column 4)
    for row in range(8, max_row + 1):
        cell = focus_ws.cell(row=row, column=4)  # Column D
        if isinstance(cell.value, (int, float)):
            # Apply comma style formatting
            cell.number_format = "#,##0"

    # Bold cells with the word 'Total' in column C
    for row in range(8, max_row + 1):
        cell = focus_ws.cell(row=row, column=3)  # Column C
        if cell.value and "total" in str(cell.value).lower():  # Check if 'total' is in the cell
            cell.font = Font(bold=True)  # Apply bold font


def apply_focus_summary_formatting(focus_ws, max_row):
    # Set headers for the Focus sheet summary section
    focus_ws["I7"].value = "Focus"  # Header for column I
    focus_ws["J7"].value = ""       # Blank column J
    focus_ws["K7"].value = "Amount" # Header for column K
    
    # Fill columns I to K in the Focus sheet with black color for headers
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    focus_ws["I7"].fill = black_fill
    focus_ws["J7"].fill = black_fill
    focus_ws["K7"].fill = black_fill
    
    # Set the font color to white for header cells
    white_font = Font(color="FFFFFF")
    focus_ws["I7"].font = white_font
    focus_ws["J7"].font = white_font
    focus_ws["K7"].font = white_font

    # Delete column J and shift everything to the left
    focus_ws.delete_cols(10)  # This deletes column J (which is the 10th column)
    
    # Round values in the new column J (which is now column K) starting from row 8
    for row in range(8, max_row + 1):
        cell = focus_ws.cell(row=row, column=9)  # Column J is now column 9 after deletion
        if isinstance(cell.value, (int, float)):  # Ensure it's a numeric value
            cell.value = round(cell.value, 0)  # Round to 0 decimal places
    
    # Apply comma style formatting to the new column J (which is now column K) starting from row 8
    for row in range(8, max_row + 1):
        cell = focus_ws.cell(row=row, column=9)  # Column J is now column 9 after deletion
        if isinstance(cell.value, (int, float)):  # Ensure it's a numeric value
            cell.number_format = '#,##0'  # Apply comma style formatting


def create_summary(focus_ws, max_row):
    summary_row = 8  # Starting row for the summary
    
    # Loop through the Focus sheet to find and summarize subtotals
    for row in range(8, max_row + 1):
        c_value = focus_ws.cell(row=row, column=3).value  # Value in column C
        d_value = focus_ws.cell(row=row, column=4).value  # Value in column D

        # Check if the value in column C contains the word "Total"
        if c_value and "Total" in str(c_value):
            # Split the value in column C into number ID and "Total"
            focus_ws.cell(row=summary_row, column=9).value = str(c_value).replace("Total", "").strip()  # Number ID in column I
            focus_ws.cell(row=summary_row, column=10).value = "Total"  # Word "Total" in column J
            focus_ws.cell(row=summary_row, column=11).value = d_value  # Subtotal value in column K
            summary_row += 1  # Move to the next row for the next summary item






def apply_subtotals_for_sheet(ws, max_row):
    current_value = None  # Tracks the current group in column C
    total_sum = 0  # Tracks the sum of the values in column D for the current group
    last_row = max_row  # The last row in the sheet

    row_idx = 8  # Start from row 8 as per your description

    while row_idx <= max_row:
        c_value = ws.cell(row=row_idx, column=3).value  # Get the value in column C
        d_value = ws.cell(row=row_idx, column=4).value  # Get the value in column D

        # If the value in column C is empty, skip this row
        if c_value is None or c_value == "":
            row_idx += 1
            continue

        # If the current value in column C is different from the previous one, insert a subtotal row
        if c_value != current_value:
            # If we have already encountered a group, insert the total row
            if current_value is not None:
                ws.insert_rows(row_idx)
                ws.cell(row=row_idx, column=3).value = f"{current_value} Total"  # Insert "Total" in column C
                ws.cell(row=row_idx, column=4).value = total_sum  # Insert the sum in column D

                # Move the row index down because we just inserted a new row
                row_idx += 1

            # Reset for the new group
            current_value = c_value
            total_sum = d_value if isinstance(d_value, (int, float)) else 0  # Start sum with first value of the new group
        else:
            # Add the current value in column D to the running total
            if isinstance(d_value, (int, float)):
                total_sum += d_value

        row_idx += 1  # Move to the next row

    # Handle the last group after the loop ends (ensure the final group is processed)
    if current_value is not None:
        ws.insert_rows(last_row + 1)  # Add the subtotal row at the end
        ws.cell(row=last_row + 1, column=3).value = f"{current_value} Total"
        ws.cell(row=last_row + 1, column=4).value = total_sum


            
def sort_focus_sheet(focus_ws, max_row):
    # Step 1: Delete rows with empty values in column C
    for row in range(max_row, 7, -1):  # Start from the bottom to avoid skipping rows
        cell = focus_ws.cell(row=row, column=3)
        if cell.value is None or cell.value == "":  # If the cell in column C is empty
            focus_ws.delete_rows(row)  # Delete the entire row

    # Step 2: Create a list to hold rows with their corresponding values in column C
    rows = []

    # Step 3: Loop through column C starting from row 8
    for row in range(8, max_row + 1):
        cell = focus_ws.cell(row=row, column=3)
        if cell.value is not None:
            c_value = str(cell.value).strip()
            
            # Append the entire row with its value in column C
            rows.append((row, c_value, [focus_ws.cell(row=row, column=col).value for col in range(1, focus_ws.max_column + 1)]))

    # Step 4: Sort rows based on the value in column C (ascending)
    # First, sort numeric values (no letters), then alphanumeric
    rows.sort(key=lambda x: (int(x[1]) if x[1].isdigit() else float('inf'), x[1]))

    # Step 5: Clear the existing values in the Focus sheet starting from row 8
    for row in range(8, max_row + 1):
        for col in range(1, focus_ws.max_column + 1):
            focus_ws.cell(row=row, column=col).value = None

    # Step 6: Write the sorted rows back into the Focus sheet
    for idx, (original_row, _, row_values) in enumerate(rows, start=8):
        for col_idx, value in enumerate(row_values, start=1):
            focus_ws.cell(row=idx, column=col_idx).value = value

def secondary_sort_focus_sheet(focus_ws, max_row):
    # Step 1: Create a list to hold rows with their corresponding values from columns C and D
    rows = []

    # Loop through column C starting from row 8 to max_row
    for row in range(8, max_row + 1):
        c_value = focus_ws.cell(row=row, column=3).value  # Column C value
        d_value = focus_ws.cell(row=row, column=4).value  # Column D value
        
        # Skip rows where column C is empty
        if c_value is None or c_value == "":
            continue
        
        # Convert column C values from text to numbers if possible
        try:
            c_value = float(c_value)
        except ValueError:
            c_value = float('-inf')  # If conversion fails, treat as the lowest possible value
        
        # Ensure column D is treated as a numeric value if possible
        if isinstance(d_value, (int, float)):
            d_value = float(d_value)  # Ensure the value is treated as a float
        else:
            d_value = float('-inf')  # Non-numeric values will be treated as the lowest possible
        
        # Append the entire row along with values from column C and D
        rows.append((row, c_value, d_value, [focus_ws.cell(row=row, column=col).value for col in range(1, focus_ws.max_column + 1)]))

    # Step 2: Sort rows based on column C (ascending) and column D (descending for same values in C)
    rows.sort(key=lambda x: (x[1], -x[2]))  # Sort by Column C ascending, then by Column D descending

    # Step 3: Clear the existing values in the sheet starting from row 8
    for row in range(8, max_row + 1):
        for col in range(1, focus_ws.max_column + 1):
            focus_ws.cell(row=row, column=col).value = None

    # Step 4: Write the sorted rows back into the sheet starting from row 8
    new_row_idx = 8
    for _, _, _, row_values in rows:
        for col_idx, value in enumerate(row_values, start=1):
            focus_ws.cell(row=new_row_idx, column=col_idx).value = value
        new_row_idx += 1







def balance_focus_grouping(file_bytes):
    # Ensure file_bytes is a BytesIO object
    if not isinstance(file_bytes, BytesIO):
        # If file_bytes is raw data (not a BytesIO object), wrap it into BytesIO
        file_bytes = BytesIO(file_bytes)

    # Load the workbook from the BytesIO object
    wb = load_workbook(filename=file_bytes)
    ws = wb.active  # Get the active worksheet

    # Create Focus worksheet
    focus_ws = wb.create_sheet(title="Focus")

    # Find the last used row and column in the original sheet
    max_row = ws.max_row
    max_col = ws.max_column

    # Step 1: Copy Column A from the original sheet into Focus sheet (Column A in Focus)
    for row in range(1, max_row + 1):
        focus_ws.cell(row=row, column=1).value = ws.cell(row=row, column=1).value

    # Step 2: Split Column A by the opening parenthesis and move to Column B and D
    for row in range(1, max_row + 1):
        val = focus_ws.cell(row=row, column=1).value
        if val and '(' in str(val):
            # Split the value at the opening parenthesis and store the parts
            parts = str(val).split('(', 1)
            focus_ws.cell(row=row, column=1).value = parts[0].strip()  # First part goes to Column A
            focus_ws.cell(row=row, column=2).value = parts[1].strip()  # Second part goes to Column D

    # Step 3: Remove parentheses in Column D (Focus sheet)
    for row in range(1, max_row + 1):
        val_d = focus_ws.cell(row=row, column=2).value
        if val_d:
            focus_ws.cell(row=row, column=2).value = str(val_d).replace("(", "").replace(")", "")
    # Step 4: Copy Column B from the original sheet to Column D in Focus sheet (as value only)
    for row in range(1, max_row + 1):
        # Get the calculated value (not the formula) from Column B in the original sheet
        original_value = ws.cell(row=row, column=2).value  # Get the value, not the formula
        focus_ws.cell(row=row, column=4).value = original_value  # Directly paste the value into Column D

    # Insert two columns at the beginning of the Focus sheet (Columns A and B become empty)
    focus_ws.insert_cols(1, 2)
    
    # Shift data from columns starting from row 5 (move the data from columns C, D, F)
    for row in range(5, max_row + 1):
        # Move data in Column C to Column E
        focus_ws.cell(row=row, column=5).value = focus_ws.cell(row=row, column=3).value
        focus_ws.cell(row=row, column=3).value = None  # Clear original cell
    
        # Move data in Column D to Column C
        focus_ws.cell(row=row, column=3).value = focus_ws.cell(row=row, column=4).value
        focus_ws.cell(row=row, column=4).value = None  # Clear original cell
    
        # Move data in Column F to Column D
        focus_ws.cell(row=row, column=4).value = focus_ws.cell(row=row, column=6).value
        focus_ws.cell(row=row, column=6).value = None  # Clear original cell
    
    # Add column titles in row 4 (ensure this is done after the rows are shifted)
    focus_ws["C4"] = "Focus"
    focus_ws["D4"] = "Amount"
    focus_ws["E4"] = "Description"
    focus_ws["F4"] = "Totals"
    
    # Shift everything below row 4 down by 3 rows in Focus sheet
    focus_ws.insert_rows(4, amount=3)  # Insert 3 rows at row 4 in focus_ws
    
    # Define fill color (black) and font color (white)
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_font = Font(color="FFFFFF")
    
    # Fill columns C to F in row 7 with black and change the text color to white in the Focus sheet
    for col in ["C", "D", "E", "F"]:
        focus_ws[f"{col}7"].fill = black_fill
        focus_ws[f"{col}7"].font = white_font
    
    # Format column D and F in the Focus sheet to show numbers with thousand commas
    for row in range(1, max_row + 1):
        focus_ws.cell(row=row, column=4).number_format = "#,##0"  # Column D
        focus_ws.cell(row=row, column=6).number_format = "#,##0"  # Column F
    
    # Increase the width of column E to double the default width in the Focus sheet
    focus_ws.column_dimensions["E"].width = focus_ws.column_dimensions["E"].width * 2.5
    
    # Call the sort_focus_sheet function after the rest of the operations in the macro
    sort_focus_sheet(focus_ws, max_row)

    # After sorting by Column C (primary sort)
    secondary_sort_focus_sheet(focus_ws, max_row=max_row)

    




    #subtotals
    apply_subtotals_for_sheet(focus_ws, max_row)
    
        
    # After sorting and performing other operations, call this function to move the last "Total" row
    move_last_total_below_group(focus_ws, start_row=8, end_row=100)

    
    create_summary(focus_ws, max_row)
    apply_focus_summary_formatting(focus_ws, max_row)


    apply_random_formatting(focus_ws, max_row)


    # Example usage
    # Assuming `focus_ws` is your worksheet object (from openpyxl)
    apply_color_coding(focus_ws)

    # Assuming `focus_ws` is your worksheet object
    calculate_totals(focus_ws, start_row=8, end_row=100)
    # Call the function to apply the comma format to Column F
    # After running your totals calculation, apply comma formatting to Column F
    apply_comma_format_no_decimal(focus_ws, start_row=8, end_row=100)



    # Save the modified workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Move cursor to the beginning of the BytesIO object
    return output.read()  # Return the transformed file as bytes





# Step 1: Upload the file
st.title("📂 Upload Your Excel File for Balance Sheet Transformation")
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    file_bytes = uploaded_file.read()

    # Process the file through the Balance Focus Grouping function
    transformed_file = balance_focus_grouping(file_bytes)

    # Store the transformed file in session state
    st.session_state.excel_bytes = transformed_file

    st.success("File processed. The 'Focus' sheet has been created and updated.")

    # Provide option to download the transformed file
    st.download_button(
        label="Download Transformed File",
        data=st.session_state.excel_bytes,
        file_name="balance_transformed_file.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
